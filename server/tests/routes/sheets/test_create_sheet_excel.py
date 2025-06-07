import json
from datetime import datetime
from io import BytesIO
from unittest.mock import patch

import pytest
from fastapi.testclient import TestClient
from openpyxl import Workbook
from tests.base_excel_test import BaseExcelTest


class CreateSheetExcelTester(BaseExcelTest):
    """Excel tester for create_sheet API (not a pytest test class)"""

    def __init__(self, app_or_client):
        """Initialize with app or client"""
        super().__init__("create_sheet_test_data")
        if isinstance(app_or_client, TestClient):
            self.client = app_or_client
        else:  # Assuming it's the FastAPI app instance
            self.client = TestClient(app_or_client)

    def execute_test_case(self, test_case: dict) -> dict:
        """
        Execute a single test case for create_sheet API

        Args:
            test_case: Dictionary containing test case data from Excel

        Returns:
            Dictionary containing test result with columns matching the desired output
        """
        try:
            # Extract test data
            sheet_name = test_case.get("SheetName", "")
            sheet_description = test_case.get("SheetDescription", "")
            column_name = test_case.get("ColumnName", "")
            column_type = test_case.get("ColumnType", "")
            column_description = test_case.get("ColumnDescription", "")

            # Create column configuration
            column_config = [
                {
                    "column_name": column_name,
                    "column_type": column_type,
                    "description": column_description,
                }
            ]

            # Create test file content
            test_file_content = self._create_test_excel_file()

            # Prepare form data for API call
            form_data = {
                "name": sheet_name,
                "description": sheet_description,
                "column_config": json.dumps(column_config),
                "status": "published",
            }

            # Create file for upload
            files = {
                "file": (
                    "test.xlsx",
                    BytesIO(test_file_content),
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            }

            # Mock the service call
            with patch("app.services.sheet_service.insert_sheet") as mock_insert, patch(
                "app.utils.asyncio_utils.run_background"
            ) as mock_background:

                mock_insert.return_value = {"id": "test-sheet-id"}

                # Make API call using self.client
                response = self.client.post("/sheets", data=form_data, files=files)

                # Determine logical test result based on validation expectations
                logical_test_result = self._determine_test_result(
                    test_case, response.status_code
                )

                # Create result record
                result_record = test_case.copy()
                result_record["ActualResult"] = logical_test_result

                # Determine 'Status' (error details or None)
                status_detail_value = None
                # Assuming 201 is the primary success code for sheet creation
                if response.status_code != 201:
                    try:
                        error_data = response.json()
                        error_detail_content = error_data.get("detail", "Unknown error")
                        if isinstance(
                            error_detail_content, list
                        ):  # Handle Pydantic error list
                            error_detail_content = json.dumps(error_detail_content)
                        status_detail_value = str(error_detail_content)
                    except json.JSONDecodeError:
                        status_detail_value = (
                            f"HTTP {response.status_code}, Body: {response.text}"
                        )
                    except (
                        Exception
                    ) as e_json:  # Catch other errors during json parsing
                        status_detail_value = f"HTTP {response.status_code}, Error parsing JSON: {str(e_json)}, Body: {response.text}"
                result_record["Status"] = status_detail_value

                result_record["Time"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

            return result_record

        except Exception as e:
            # This handles exceptions during the test execution itself (e.g., API call failure)
            failed_result = test_case.copy()
            # Mark as Fail if test execution itself errors
            failed_result["ActualResult"] = "Fail"
            failed_result["Status"] = f"Test execution error: {str(e)}"
            failed_result["Time"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            # Ensure all expected columns are present even in error cases
            for key in [
                "SheetName",
                "SheetDescription",
                "ColumnName",
                "ColumnType",
                "ColumnDescription",
                "ExpectedResult",
            ]:
                if key not in failed_result:
                    failed_result[key] = "N/A due to error"
            return failed_result

    def _create_test_excel_file(self) -> bytes:
        """Create a simple Excel file for testing"""
        wb = Workbook()
        ws = wb.active
        ws["A1"] = "Test Data"

        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer.getvalue()

    def _determine_test_result(self, test_case: dict, status_code: int) -> str:
        """
        Ghi nhận trạng thái thực tế của API:
        - Nếu status_code == 201: ActualResult = 'Pass'
        - Nếu status_code != 201: ActualResult = 'Fail'
        """
        return "Pass" if status_code == 201 else "Fail"


@pytest.mark.asyncio
async def test_create_sheet_from_excel(test_app):
    """Test create_sheet API using Excel test data"""
    tester = CreateSheetExcelTester(test_app)  # Create tester with test_app

    # Run all tests from Excel file
    results = tester.run_all_tests()  # This method is synchronous

    # Get test summary
    summary = tester.get_test_summary(results)

    print(f"\nTest Summary for create_sheet API:")
    print(f"Total tests: {summary['total']}")
    print(f"Passed: {summary['passed']}")
    print(f"Failed: {summary['failed']}")
    print(f"Pass rate: {summary['pass_rate']}")
    print(f"Results written to: {tester.test_file_path}")

    # Assert that we have some tests and at least some should pass
    assert len(results) > 0, "No test cases found"
    # This assertion might be too strict if all valid cases are designed to pass
    # and all invalid cases are designed to fail (and be caught as 'Pass' by _determine_test_result)
    # A better assertion might be that summary['failed'] == 0 if all tests behave as expected by _determine_test_result
    assert (
        summary["failed"] == 0
    ), f"{summary['failed']} tests did not match expected outcomes. Check {tester.test_file_path}"


if __name__ == "__main__":
    import asyncio
    import os
    import sys

    # --- Setup for direct run ---
    # Add project root (server directory) to sys.path to allow 'from app.main import app'
    # This assumes the script is in server/tests/routes/sheets/
    current_dir = os.path.dirname(os.path.abspath(__file__))
    server_root = os.path.abspath(os.path.join(current_dir, "..", "..", ".."))
    if server_root not in sys.path:
        sys.path.insert(0, server_root)

    from app.main import app as main_fastapi_app  # Import for __main__

    async def main():
        """Run the tests directly"""
        # Create test client
        # Create and run tester
        test_client = TestClient(main_fastapi_app)
        tester = CreateSheetExcelTester(test_client)
        results = tester.run_all_tests()

        # Print summary
        summary = tester.get_test_summary(results)
        print(f"\nTest Summary for create_sheet API:")
        print(f"Total tests: {summary['total']}")
        print(f"Passed: {summary['passed']}")
        print(f"Failed: {summary['failed']}")
        print(f"Pass rate: {summary['pass_rate']}")
        print(f"Results written to: {tester.test_file_path}")

    # Run directly
    asyncio.run(main())
