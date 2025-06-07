"""
Base class for Excel-based testing
"""

import os
from abc import ABC, abstractmethod
from typing import Dict, List

import pandas as pd
from openpyxl import load_workbook


class BaseExcelTest(ABC):
    """Base class for all Excel-based tests"""

    def __init__(self, test_file_name: str):
        """
        Initialize the test with Excel file

        Args:
            test_file_name: Name of the Excel file (without .xlsx extension)
        """
        self.test_file_name = test_file_name
        self.test_file_path = self._get_test_file_path()

    def _get_test_file_path(self) -> str:
        """Get the full path to the test Excel file"""
        return os.path.join(
            os.path.dirname(__file__), "data", "excel", f"{self.test_file_name}.xlsx"
        )

    def read_test_data(self, sheet_name: str = "test-data") -> List[Dict]:
        """
        Read test data from Excel sheet

        Args:
            sheet_name: Name of the sheet to read from

        Returns:
            List of dictionaries containing test data
        """
        if not os.path.exists(self.test_file_path):
            raise FileNotFoundError(f"Test file not found: {self.test_file_path}")

        df = pd.read_excel(self.test_file_path, sheet_name=sheet_name)
        # Convert NaN to None and return as list of dicts
        return df.where(pd.notnull(df), None).to_dict("records")

    def write_test_results(self, results: List[Dict], sheet_name: str = "test-result"):
        """
        Write test results to Excel sheet

        Args:
            results: List of dictionaries containing test results
            sheet_name: Name of the sheet to write to
        """
        if not results:
            return

        # No longer adding timestamp here, it's added in execute_test_case

        wb = load_workbook(self.test_file_path)

        # Check if result sheet exists, if not create it
        if sheet_name not in wb.sheetnames:
            ws = wb.create_sheet(sheet_name)
            # Write headers
            headers = list(results[0].keys())
            for col, header in enumerate(headers, 1):
                ws.cell(row=1, column=col, value=header)
        else:
            ws = wb[sheet_name]
            # Clear existing data (except headers)
            ws.delete_rows(2, ws.max_row)

        # Write results
        for row, result in enumerate(results, 2):
            for col, (key, value) in enumerate(result.items(), 1):
                ws.cell(row=row, column=col, value=value)

        wb.save(self.test_file_path)

    @abstractmethod
    def execute_test_case(self, test_case: Dict) -> Dict:
        """
        Execute a single test case

        Args:
            test_case: Dictionary containing test case data

        Returns:
            Dictionary containing test result with 'Result' field
        """

    def run_all_tests(self, sheet_name: str = "test-data") -> List[Dict]:
        """
        Run all test cases from Excel file

        Args:
            sheet_name: Name of the sheet containing test data

        Returns:
            List of test results
        """
        test_data = self.read_test_data(sheet_name)
        results = []

        for test_case in test_data:
            try:
                result = self.execute_test_case(test_case)
                results.append(result)
            except Exception as e:
                # Add failed result if execution throws exception
                failed_result = test_case.copy()
                failed_result["Result"] = "Fail"
                failed_result["ErrorMessage"] = str(e)
                results.append(failed_result)

        # Write results back to Excel
        self.write_test_results(results)

        return results

    def get_test_summary(self, results: List[Dict]) -> Dict:
        """
        Get summary of test results

        Args:
            results: List of test results

        Returns:
            Dictionary with test summary
        """
        total = len(results)
        # A test passes if:
        # - ExpectedResult=Pass and ActualResult=Pass
        # - ExpectedResult=Fail and ActualResult=Pass (API correctly rejected the request)
        # In other words, ActualResult should always be 'Pass' if our API is working correctly
        # Một test chỉ pass nếu ExpectedResult == ActualResult
        passed = sum(
            1
            for r in results
            if str(r.get("ExpectedResult", "")).strip().lower()
            == str(r.get("ActualResult", "")).strip().lower()
        )
        failed = total - passed

        return {
            "total": total,
            "passed": passed,
            "failed": failed,
            "pass_rate": f"{(passed/total*100):.1f}%" if total > 0 else "0%",
        }
